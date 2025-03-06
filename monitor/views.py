from django.shortcuts import render, redirect
from .forms import UserInputForm, QuoteForm
from .models import UserInput, Quote
from django.utils import timezone
from django.db.models import Sum
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


@login_required
def calendar_view(request):
    return render(request, 'monitor/calendar.html')

# Login View
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')  # Redirect to home page after successful login
        else:
            messages.error(request, 'Invalid username or password')  # Show error message
    
    return render(request, 'monitor/login.html')  # Render the login page


# Logout View
def logout_view(request):
    logout(request)
    return redirect('login')  # Redirect to login page after logout


# Home View (after login)
@login_required
def home_view(request):
    quotes = Quote.objects.all()  # Retrieve all quotes from the database
    form = QuoteForm()

    if request.method == "POST":
        form = QuoteForm(request.POST)
        if form.is_valid():
            form.save()  # Save the new quote to the database
            return redirect('home')  # Redirect to the home page after saving the quote

    return render(request, 'monitor/home.html', {'quotes': quotes, 'form': form})

@csrf_exempt
def delete_quote(request, id):
    if request.method == "POST":
        try:
            quote = Quote.objects.get(id=id)
            quote.delete()
            return JsonResponse({"success": True})
        except Quote.DoesNotExist:
            return JsonResponse({"success": False, "error": "Quote not found"})
    return JsonResponse({"success": False, "error": "Invalid request method"})

# Input View
@login_required  # Ensure that only authenticated users can submit data
def input_view(request):
    if request.method == 'POST':
        form = UserInputForm(request.POST)
        if form.is_valid():
            user_input = form.save(commit=False)
            user_input.date = timezone.now().date()  # Set the current date
            user_input.save()
            return redirect('home')  # Redirect to the home page after saving the form
    else:
        form = UserInputForm()

    current_date = timezone.now().date()  # Get today's date
    return render(request, 'monitor/input.html', {'form': form, 'current_date': current_date})

# Chart View
@login_required  # Ensure that only logged-in users can access the chart page
def chart_view(request):
    # Retrieve all UserInput records
    inputs = UserInput.objects.all().order_by('date')  # Order by date for the time series

    # If no inputs are available, set default values to avoid errors
    if not inputs:
        line_chart_data = {'labels': [], 'cumulative_score': []}
        bar_chart_data = {'labels': [], 'values': []}
        return render(request, 'monitor/chart.html', {
            'line_chart_data': line_chart_data,
            'bar_chart_data': bar_chart_data,
        })

    # Prepare cumulative data for the line chart
    date_labels = []
    cumulative_data = []
    current_score = 0  # Start with a score of 0

    # Prepare bar chart data variables
    total_read_bible = 0
    total_prayed = 0
    total_exercised = 0
    total_book_reading = 0
    total_studying_machine_learning = 0
    total_masturbated = 0
    total_cultivated_relationships = 0
    total_woke_up_at_5am = 0
    total_healthy_eating = 0
    total_hurt_someone = 0
    total_wasted_time = 0

    for user_input in inputs:
        date_labels.append(user_input.date.strftime('%Y-%m-%d'))  # Use the date field for x-axis labels

        # Calculate cumulative score based on user inputs with weights
        if user_input.read_bible:
            current_score += 2  # Higher weight for reading the Bible
            total_read_bible += 1
        else:
            current_score -= 3  # Negative weight for not reading the Bible

        if user_input.prayer_god:
            current_score += 1
            total_prayed += 1
        if user_input.exercised_or_played:
            current_score += 1
            total_exercised += 1
        if user_input.book_reading:
            current_score += 1
            total_book_reading += 1
        if user_input.studying_machine_learning:
            current_score += 1
            total_studying_machine_learning += 1
        if user_input.cultivated_relationships:
            current_score += 1
            total_cultivated_relationships += 1
        if user_input.woke_up_at_5am:
            current_score += 1
            total_woke_up_at_5am += 1
        if user_input.healthy_eating:
            current_score += 1
            total_healthy_eating += 1
        
        if user_input.hurt_someone:
            current_score -= 1  # Negative weight for hurting someone
            total_hurt_someone += 1
        if user_input.maintained_purity:
            current_score -= 2  # Higher negative weight for masturbation
            total_masturbated += 1
        if user_input.wasted_time:
            current_score -= 1  # Negative weight for wasting time
            total_wasted_time += 1

        cumulative_data.append(current_score)  # Append the current score to the data list

    line_chart_data = {
        'labels': date_labels,
        'cumulative_score': cumulative_data,
    }

    # Prepare bar chart data by summing individual inputs
    bar_chart_data = {
        'labels': [
            'Read Bible', 
            'Prayed', 
            'Exercised', 
            'Book Reading', 
            'Studying ML', 
            'Cultivated Relationships', 
            'Woke Up at 5 AM', 
            'Healthy Eating', 
            'Hurt Someone', 
            'Purity', 
            'Wasted Time'
        ],
        'values': [
            total_read_bible,
            total_prayed,
            total_exercised,
            total_book_reading,
            total_studying_machine_learning,
            total_cultivated_relationships,
            total_woke_up_at_5am,
            total_healthy_eating,
            -total_hurt_someone,  # Show hurt as a negative value
            -total_masturbated,  # Show masturbation as a negative value
            -total_wasted_time  # Show wasted time as a negative value
        ],
    }

    return render(request, 'monitor/chart.html', {
        'line_chart_data': line_chart_data,
        'bar_chart_data': bar_chart_data,
    })



@login_required  # Ensure only logged-in users can export the data
def export_to_excel_view(request):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Inputs"

    # Add headers to the first row of the Excel sheet
    headers = [
        'Date', 
        'Read Bible', 
        'Prayed', 
        'Exercised', 
        'Book Reading', 
        'Studying ML', 
        'Cultivated Relationships', 
        'Woke Up at 5 AM', 
        'Healthy Eating', 
        'Hurt Someone', 
        'Purity', 
        'Wasted Time', 
        'Daily Summary'  # Ensure to include Daily Summary at the end
    ]
    
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet[f'{column_letter}1'] = header

    # Retrieve all UserInput records and add them to the Excel sheet
    inputs = UserInput.objects.all().order_by('date')
    for row_num, user_input in enumerate(inputs, 2):
        sheet[f'A{row_num}'] = user_input.date.strftime('%Y-%m-%d')
        sheet[f'B{row_num}'] = 'Yes' if user_input.read_bible else 'No'
        sheet[f'C{row_num}'] = 'Yes' if user_input.prayer_god else 'No'
        sheet[f'D{row_num}'] = 'Yes' if user_input.exercised_or_played else 'No'
        sheet[f'E{row_num}'] = 'Yes' if user_input.book_reading else 'No'  # Include Book Reading field
        sheet[f'F{row_num}'] = 'Yes' if user_input.studying_machine_learning else 'No'  # Include Studying ML field
        sheet[f'G{row_num}'] = 'Yes' if user_input.cultivated_relationships else 'No'  # Include Cultivated Relationships field
        sheet[f'H{row_num}'] = 'Yes' if user_input.woke_up_at_5am else 'No'  # Include Woke Up at 5 AM field
        sheet[f'I{row_num}'] = 'Yes' if user_input.healthy_eating else 'No'  # Include Healthy Eating field
        sheet[f'J{row_num}'] = 'Yes' if user_input.hurt_someone else 'No'
        sheet[f'K{row_num}'] = 'Yes' if user_input.maintained_purity else 'No'  # Include Masturbated field
        sheet[f'L{row_num}'] = 'Yes' if user_input.wasted_time else 'No'  # Include Wasted Time field
        sheet[f'M{row_num}'] = user_input.daily_summary  # Daily Summary field

    # Create an HTTP response with the Excel file as an attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_inputs.xlsx"'
    
    # Save the workbook to the response
    workbook.save(response)
    
    return response

